"""
Real Dataset Loader for ML Model Training
Automatically loads datasets from data/ folder and preprocesses them
"""

import os
import csv
from typing import List, Dict, Optional


def load_uci_energy_data(filepath: str) -> List[Dict]:
    """
    Load UCI Energy Efficiency Dataset
    Expected columns: X1-X8 (building params), Y1-Y2 (heating/cooling loads)
    """
    data = []
    
    try:
        # Try CSV first
        if filepath.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Map UCI columns to our format
                    area = float(row.get('X2', 0)) * 40  # Wall area * multiplier
                    budget = min(100, max(20, 50 + float(row.get('X7', 0)) * 10))  # Glazing area influence
                    
                    # Infer climate from heating/cooling ratio
                    heating = float(row.get('Y1', 15))
                    cooling = float(row.get('Y2', 15))
                    if heating > cooling * 1.5:
                        climate = 'cold'
                    elif cooling > heating * 1.5:
                        climate = 'hot'
                    else:
                        climate = 'moderate'
                    
                    # Energy efficiency from heating load (inverted)
                    energy_eff = max(0, min(100, 100 - (heating * 3)))
                    
                    data.append({
                        'area': int(area),
                        'budget': int(budget),
                        'climate': climate,
                        'priority': 'energy',  # UCI focuses on energy
                        'design_id': 0,
                        'actual_cost': int(area * 150 * (budget / 100) * 1.2),
                        'energy_efficiency': int(energy_eff),
                        'water_efficiency': 65,
                        'carbon_level': 'Low' if energy_eff > 70 else 'Medium'
                    })
        
        # Try Excel if available
        elif filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    
                    area = float(row_dict.get('X2', 0)) * 40
                    budget = min(100, max(20, 50 + float(row_dict.get('X7', 0)) * 10))
                    
                    heating = float(row_dict.get('Y1', 15))
                    cooling = float(row_dict.get('Y2', 15))
                    if heating > cooling * 1.5:
                        climate = 'cold'
                    elif cooling > heating * 1.5:
                        climate = 'hot'
                    else:
                        climate = 'moderate'
                    
                    energy_eff = max(0, min(100, 100 - (heating * 3)))
                    
                    data.append({
                        'area': int(area),
                        'budget': int(budget),
                        'climate': climate,
                        'priority': 'energy',
                        'design_id': 0,
                        'actual_cost': int(area * 150 * (budget / 100) * 1.2),
                        'energy_efficiency': int(energy_eff),
                        'water_efficiency': 65,
                        'carbon_level': 'Low' if energy_eff > 70 else 'Medium'
                    })
                    
            except ImportError:
                print("⚠ openpyxl not installed. Install with: pip install openpyxl")
                return []
        
        print(f"✓ Loaded {len(data)} records from UCI Energy dataset")
        return data
        
    except Exception as e:
        print(f"⚠ Failed to load UCI data: {e}")
        return []


def load_nyc_building_data(filepath: str, max_rows: int = 5000) -> List[Dict]:
    """
    Load NYC Building Energy & Water Data
    """
    data = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            count = 0
            
            for row in reader:
                if count >= max_rows:
                    break
                
                try:
                    # Extract and validate area
                    area_str = row.get('Property GFA - Self-Reported (ft²)', '0')
                    area = float(area_str.replace(',', '')) if area_str else 0
                    if area < 300 or area > 50000:
                        continue
                    
                    # Normalize to our range
                    area = min(2000, max(300, int(area / 10)))
                    
                    # Extract energy (kBtu)
                    energy_str = row.get('Site EUI (kBtu/ft²)', '0')
                    energy_use = float(energy_str) if energy_str else 100
                    energy_eff = max(0, min(100, 100 - (energy_use / 3)))
                    
                    # Extract water (kgal)
                    water_str = row.get('Water Use (All Water Sources) (kgal)', '0')
                    water_use = float(water_str.replace(',', '')) if water_str else 0
                    water_eff = max(0, min(100, 100 - (water_use / area / 2)))
                    
                    # Extract GHG emissions
                    ghg_str = row.get('Total GHG Emissions (Metric Tons CO2e)', '0')
                    ghg = float(ghg_str.replace(',', '')) if ghg_str else 0
                    if ghg < area * 0.5:
                        carbon_level = 'Low'
                    elif ghg < area * 1.5:
                        carbon_level = 'Medium'
                    else:
                        carbon_level = 'High'
                    
                    # Infer climate from location (NYC is moderate)
                    climate = 'moderate'
                    
                    # Infer priority from efficiency scores
                    if energy_eff > water_eff:
                        priority = 'energy'
                    elif water_eff > energy_eff:
                        priority = 'water'
                    else:
                        priority = 'materials'
                    
                    # Budget from energy star score
                    score_str = row.get('ENERGY STAR Score', '50')
                    score = float(score_str) if score_str else 50
                    budget = min(100, max(20, int(score)))
                    
                    data.append({
                        'area': area,
                        'budget': budget,
                        'climate': climate,
                        'priority': priority,
                        'design_id': 0 if energy_eff > 70 else 1,
                        'actual_cost': int(area * 180 * (budget / 100)),
                        'energy_efficiency': int(energy_eff),
                        'water_efficiency': int(water_eff),
                        'carbon_level': carbon_level
                    })
                    
                    count += 1
                    
                except (ValueError, KeyError, TypeError):
                    continue
        
        print(f"✓ Loaded {len(data)} records from NYC Building dataset")
        return data
        
    except Exception as e:
        print(f"⚠ Failed to load NYC data: {e}")
        return []


def load_custom_csv(filepath: str) -> List[Dict]:
    """
    Load custom CSV with our exact format OR UCI Appliances Energy format
    """
    data = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Try custom format first
                if 'area' in row:
                    data.append({
                        'area': int(row['area']),
                        'budget': int(row['budget']),
                        'climate': row['climate'],
                        'priority': row['priority'],
                        'design_id': int(row['design_id']),
                        'actual_cost': int(row['actual_cost']),
                        'energy_efficiency': int(row.get('energy_efficiency', 50)),
                        'water_efficiency': int(row.get('water_efficiency', 50)),
                        'carbon_level': row.get('carbon_level', 'Medium')
                    })
                # Try UCI Appliances Energy format
                elif 'Appliances' in row:
                    try:
                        appliances = float(row.get('Appliances', 0).strip())
                        lights = float(row.get('lights', 0).strip())
                        t_out = float(row.get('T_out', 20).strip())
                        
                        # Map to our format
                        area = int(500 + (appliances + lights) * 2)  # Estimate area
                        energy_eff = max(0, min(100, 100 - (appliances / 100)))
                        
                        climate = 'cold' if t_out < 5 else ('hot' if t_out > 20 else 'moderate')
                        
                        data.append({
                            'area': area,
                            'budget': int(50 + (energy_eff / 100) * 50),
                            'climate': climate,
                            'priority': 'energy',
                            'design_id': 0,
                            'actual_cost': int(area * 150),
                            'energy_efficiency': int(energy_eff),
                            'water_efficiency': 65,
                            'carbon_level': 'Low' if energy_eff > 70 else 'Medium'
                        })
                    except (ValueError, KeyError):
                        continue
        
        print(f"✓ Loaded {len(data)} records from custom CSV")
        return data
        
    except Exception as e:
        print(f"⚠ Failed to load custom data: {e}")
        return []


def auto_load_training_data(data_dir: str = 'data') -> Dict[str, List[Dict]]:
    """
    Automatically discover and load all available datasets
    Returns dict with 'cost', 'preference', and 'historical' data
    """
    result = {
        'cost': [],
        'preference': [],
        'historical': []
    }
    
    if not os.path.exists(data_dir):
        print(f"⚠ Data directory not found: {data_dir}")
        return result
    
    print(f"🔍 Scanning {data_dir} for training data...")
    
    # Try UCI Energy Efficiency
    for filename in ['energy_efficiency.xlsx', 'energy_efficiency.csv', 'ENB2012_data.xlsx']:
        filepath = os.path.join(data_dir, filename)
        if os.path.exists(filepath):
            uci_data = load_uci_energy_data(filepath)
            if uci_data:
                result['cost'].extend(uci_data)
                result['preference'].extend(uci_data)
                result['historical'].extend(uci_data)
            break
    
    # Try NYC Building Data
    for filename in ['NYC.csv', 'nyc_building_energy.csv', 'nyc_buildings.csv']:
        filepath = os.path.join(data_dir, filename)
        if os.path.exists(filepath):
            nyc_data = load_nyc_building_data(filepath, max_rows=10000)
            if nyc_data:
                result['cost'].extend(nyc_data)
                result['preference'].extend(nyc_data)
                result['historical'].extend(nyc_data)
            break
    
    # Try custom data
    for filename in ['energy_data.csv', 'custom_data.csv', 'training_data.csv', 'sample_training_data.csv']:
        filepath = os.path.join(data_dir, filename)
        if os.path.exists(filepath):
            custom_data = load_custom_csv(filepath)
            if custom_data:
                result['cost'].extend(custom_data)
                result['preference'].extend(custom_data)
                result['historical'].extend(custom_data)
    
    # Load synthetic training data (high-quality generated dataset)
    synthetic_path = os.path.join(data_dir, 'synthetic_training_data.csv')
    if os.path.exists(synthetic_path):
        synthetic_data = load_custom_csv(synthetic_path)
        if synthetic_data:
            result['cost'].extend(synthetic_data)
            result['preference'].extend(synthetic_data)
            result['historical'].extend(synthetic_data)
            print(f"✓ Loaded {len(synthetic_data)} records from synthetic training data")
    
    # Summary
    if result['cost']:
        print(f"✓ Total training data loaded: {len(result['cost'])} samples (ENHANCED with synthetic data)")
    else:
        print("ℹ No real datasets found. Will use synthetic data.")
    
    return result


def prepare_cost_training_data(raw_data: List[Dict]) -> List[Dict]:
    """Convert raw data to cost predictor format"""
    return [{
        'area': d['area'],
        'budget': d['budget'],
        'climate': d['climate'],
        'priority': d['priority'],
        'design_id': d['design_id'],
        'actual_cost': d['actual_cost']
    } for d in raw_data]


def prepare_preference_training_data(raw_data: List[Dict]) -> List[Dict]:
    """Convert raw data to design ranker format"""
    prepared = []
    
    for d in raw_data:
        prepared.append({
            'constraints': {
                'area': d['area'],
                'budget': d['budget'],
                'climate': d['climate'],
                'priority': d['priority']
            },
            'designs': [{
                'id': 'design-a',
                'metrics': {
                    'energyEfficiency': d.get('energy_efficiency', 50),
                    'waterEfficiency': d.get('water_efficiency', 50),
                    'carbonFootprint': d.get('carbon_level', 'Medium'),
                    'estimatedCost': d['actual_cost']
                }
            }],
            'ranking': [0],
            'satisfaction': 0.85 if d.get('energy_efficiency', 50) > 70 else 0.65
        })
    
    return prepared


def prepare_historical_training_data(raw_data: List[Dict]) -> List[Dict]:
    """Convert raw data to recommender format"""
    return [{
        'constraints': {
            'area': d['area'],
            'budget': d['budget'],
            'climate': d['climate'],
            'priority': d['priority']
        },
        'chosen_design': d['design_id'],
        'satisfaction': 0.85 if d.get('energy_efficiency', 50) > 70 else 0.70
    } for d in raw_data]
